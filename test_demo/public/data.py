import csv
import random, os, re
import pandas
import string
from random import shuffle
from datetime import datetime, timedelta
from dateutil.relativedelta import relativedelta
import openpyxl

filename = r'C:\\Users\\19144\\PycharmProjects\\学习\\test_demo\\data\\userinfo.csv'
filename1 = r'C:\\Users\\19144\\PycharmProjects\\学习\\test_demo\\data\\url.txt'
filename2 = r'C:\\Users\\19144\\PycharmProjects\\学习\\test_demo\\data\\userimf.csv'
filename3 = r'C:\\Users\\19144\\Documents\\xiufu\\jmeter\\'


# 获取data文件下userinfo.csv里面的用户名和密码
def get_un_pw():
    # 读取本地 CSV 文件
    with open(filename, encoding='utf-8') as f:
        row = csv.reader(f, delimiter=',')
        next(row)  # 从第二行开始读取
        # next(row)
        height = []
        for r in row:
            height.append(r[0])  # 读取的字段均为str类型
            height.append(r[1])
    return height


# 获取一个密码列表，最后一项为随机正常格式的密码
def get_new_passwrod():
    ret_list_passwds = [";3ss", "311h1kjjx.w;1W3849:h2j1h2", "22231212", "ahsdndkflf", ",;*.;$,;."]
    src = string.ascii_letters + string.digits
    list_passwd_all = random.sample(src, 5)  # 从字母和数字中随机取5位
    list_passwd_all.extend(random.sample(string.digits, 1))  # 让密码中一定包含数字
    list_passwd_all.extend(random.sample(string.ascii_lowercase, 1))  # 让密码中一定包含小写字母
    list_passwd_all.extend(random.sample(string.ascii_uppercase, 1))  # 让密码中一定包含大写字母
    random.shuffle(list_passwd_all)  # 打乱列表顺序
    str_passwd = ''.join(list_passwd_all)  # 将列表转化为字符串
    ret_list_passwds.append(str_passwd)
    return ret_list_passwds


# 修改data文件下userinfo.csv里面的密码
def change_passwrod(str):
    df = pandas.read_csv(filename, header=0, index_col=0, encoding='utf-8')
    # read_csv的教程：https://www.jianshu.com/p/ebb64a159104
    df["密码"][0] = str
    df.to_csv(filename, encoding='utf-8')
    # df.close_csv()


def write_imf(name, mail, phonenum, psd):
    df = pandas.read_csv(filename2, header=0, index_col=0, encoding='utf-8')
    # read_csv的教程：https://www.jianshu.com/p/ebb64a159104
    df.loc[len(df)] = [name, mail, phonenum, psd]  # loc[n]表示将第n行替换为后面的列表
    df.to_csv(filename2, encoding='utf-8')
    # df.close_csv()


# 获取data文件下url.txt里面第一行url
def get_url():
    f = open(filename1, "r")
    url = f.readline()
    f.close()
    return url


def GBK2312():  # 供get_some_name使用，生成一个GBK2312编码的字符，含有的生僻字比unicode编码字符少一些
    head = random.randint(0xb0, 0xf7)
    body = random.randint(0xa1, 0xf9)  # 在head区号为55的那一块最后5个汉字是乱码,为了方便缩减下范围
    val = f'{head:x}{body:x}'
    str = bytes.fromhex(val).decode('gb2312')
    return str


# 获取一个列表，有各种名称
def get_some_name():
    salt = ".!@#$%^&*()><?;/,[]{}|"
    ret = []
    # 单个汉字
    # val = random.randint(0x4e00, 0x9fbf)
    # ret.append(str(chr(val)))
    ret.append(GBK2312())
    # 单个数字
    val = random.randint(0, 9)
    ret.append(str(val))
    # 单个字母
    val = random.choice(string.ascii_letters)
    ret.append(val)
    # 单个字符
    val = random.choice(salt)
    ret.append(val)
    # 13位汉字、12位数字、11位字母
    cout = 13
    zh = ''
    num = ''
    en = ''
    while (cout > 0):
        # zh = zh + str(chr(random.randint(0x4e00, 0x8fbf)))
        zh = zh + GBK2312()
        if cout == 2:
            cout = cout - 1
            continue
        num = num + str(random.randint(0, 9))
        if cout == 1:
            cout = cout - 1
            continue
        en = en + random.choice(string.ascii_letters)
        cout = cout - 1
    ret.append(zh)
    ret.append(num)
    ret.append(en)
    # 6位 汉字+字符 数字+字符 字母+字符
    cout = 3
    zh = []
    num = []
    en = []
    while (cout > 0):
        # zh.append(chr(random.randint(0x4e00, 0x8fbf)))
        zh.append(GBK2312())
        zh.append(random.choice(salt))
        num.append(str(random.randint(0, 9)))
        num.append(random.choice(salt))
        en.append(random.choice(string.ascii_letters))
        en.append(random.choice(salt))
        cout = cout - 1
    shuffle(zh)
    shuffle(num)
    shuffle(en)
    ret.append(''.join(zh))
    ret.append(''.join(num))
    ret.append(''.join(en))
    # 2位汉字，5位数字，10位字母
    cout = 10
    zh = ''
    num = ''
    en = ''
    while (cout > 0):
        en = en + random.choice(string.ascii_letters)
        if (cout < 6):
            cout = cout - 1
            continue
        num = num + str(random.randint(0, 9))
        if (cout < 9):
            cout = cout - 1
            continue
        # zh = zh + str(chr(random.randint(0x4e00, 0x8fbf)))
        zh = zh + GBK2312()
        cout = cout - 1
    ret.append(zh)
    ret.append(num)
    ret.append(en)
    # 6位 汉字数字字母混合
    cout = 2
    zne = []
    while (cout > 0):
        # zne.append(str(chr(random.randint(0x4e00, 0x8fbf))))
        zne.append(GBK2312())
        zne.append(str(random.randint(0, 9)))
        zne.append(random.choice(string.ascii_letters))
        cout = cout - 1
    shuffle(zne)
    ret.append(''.join(zne))
    print(ret)
    return ret


# 获取一个列表，有3个邮箱地址
def get_some_mailbox():
    ret = []
    j = 3
    while (j > 0):
        i = 9
        nstr = ''
        while (i > 0):
            val = random.randint(0, 9)
            if (i == 9 and val == 0):
                continue
            nstr = nstr + str(val)
            i = i - 1
        ret.append(nstr + "@qq.com")
        j = j - 1
    return ret


# 获取一个列表，有3个手机号
def get_some_phonenum():
    ret = []
    j = 3
    while (j > 0):
        i = 9
        nstr = random.choice(["13", "14", "15", "16", "17", "18", "19"])
        while (i > 0):
            val = random.randint(0, 9)
            nstr = nstr + str(val)
            i = i - 1
        ret.append(nstr)
        j = j - 1
    return ret


# 返回一个时间字符串，可相对于今天，往前和往后的任意时段
def delay_time(years=0, months=0, days=0, hours=0, minutes=0, seconds=0):
    time_str = datetime.now()
    # if type(time_str) == str:
    #     time_str = datetime.strptime(time_str, '%Y-%m-%d %H:%M:%S')
    ret = time_str + relativedelta(years=years, months=months, days=days, hours=hours, minutes=minutes, seconds=seconds)
    hh = str(ret)
    hh = hh[:-7]
    return hh


# 返回文件夹下的一个文件
def get_file_name(filepath):
    for i, j, k in os.walk(filepath):
        ret = random.choice(k)
    return ret


# 获取jameter文件夹下的用户名和密码
def get_un_pw_jameter():
    # 读取用户管理表格.xlsx文件
    wb = openpyxl.load_workbook(filename3 + "用户管理表格.xlsx")
    # 获取所有工作表名
    names = wb.sheetnames
    # wb.get_sheet_by_name(name) 已经废弃,使用wb[name] 获取指定工作表
    sheet = wb[names[0]]
    # 获取最大行数
    maxRow = sheet.max_row
    # 获取最大列数
    maxColumn = sheet.max_column
    # 获取当前活动表
    current_sheet = wb.active
    # 获取当前活动表名称
    current_name = sheet.title
    # # 通过名字访问Cell对象, 通过value属性获取值
    # a1 = sheet['A1'].value
    # # 通过行和列确定数据
    # a12 = sheet.cell(row=1, column=2).value
    # # 获取列字母
    # column_name = openpyxl.utils.cell.get_column_letter(1)
    # # 将列字母转为数字, 参数忽略大小写
    # column_name_num = openpyxl.utils.cell.column_index_from_string('a')
    # # 获取一列数据, sheet.iter_rows() 获取所有的行
    # """
    # (<Cell 'Sheet1'.A1>, <Cell 'Sheet1'.B1>, <Cell 'Sheet1'.C1>)
    # (<Cell 'Sheet1'.A2>, <Cell 'Sheet1'.B2>, <Cell 'Sheet1'.C2>)
    # (<Cell 'Sheet1'.A3>, <Cell 'Sheet1'.B3>, <Cell 'Sheet1'.C3>)
    # (<Cell 'Sheet1'.A4>, <Cell 'Sheet1'.B4>, <Cell 'Sheet1'.C4>)
    # (<Cell 'Sheet1'.A5>, <Cell 'Sheet1'.B5>, <Cell 'Sheet1'.C5>)
    # """
    # for one_column_data in sheet.iter_rows():
    #     print(one_column_data[0].value)
    #
    # # 获取一行数据, sheet.iter_cols() 获取所有的列
    # """
    # (<Cell 'Sheet1'.A1>, <Cell 'Sheet1'.A2>, <Cell 'Sheet1'.A3>)
    # (<Cell 'Sheet1'.B1>, <Cell 'Sheet1'.B2>, <Cell 'Sheet1'.B3>)
    # (<Cell 'Sheet1'.C1>, <Cell 'Sheet1'.C2>, <Cell 'Sheet1'.C3>)
    # """
    # for one_row_data in sheet.iter_cols():
    #     print(one_row_data[1].value)
    #
    # print("row = {}, column = {}".format(maxRow, maxColumn))
    ret = []
    i = 2
    while (i < 62):
        ret.append(sheet.cell(row=i, column=6).value)
        i = i + 1
    return ret


if __name__ == "__main__":
    print(get_un_pw())
